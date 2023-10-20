#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Oct 16 14:24:11 2023

@author: mj
"""
import openpyxl as pyxl
import sys
import numpy as np
from math import radians, cos, sin, asin, sqrt


file = input("Input file name with filepath \n")
startingCity = input("Input starting city as CityName, 2 Character State (ex: Cleveland, OH \n")

def getData(filepath):
    try:
        wb = pyxl.load_workbook(filepath)
        print("File successfully loaded")
    except:
        sys.exit("Failed opening file")
    
    ws = wb['Sheet1']
    currentCell = 0
    cityList = []
    rowNum = 2
    while currentCell is not None:
        currentCell = ws.cell(row = rowNum, column = 1).value
        currentLat = ws.cell(row = rowNum, column = 2).value
        currentLon = ws.cell(row = rowNum, column = 3).value
        cityList += [[currentCell, currentLat, currentLon]]
        rowNum += 1
    #remove the last none value from cityList
    cityList.pop(-1)
    
    return cityList

cityList = getData(file)

#thank you geeks4geeks
#https://www.geeksforgeeks.org/program-distance-two-points-earth/#
def distance(lat1, lat2, lon1, lon2):
     
    # The math module contains a function named
    # radians which converts from degrees to radians.
    lon1 = radians(lon1)
    lon2 = radians(lon2)
    lat1 = radians(lat1)
    lat2 = radians(lat2)
      
    # Haversine formula 
    dlon = lon2 - lon1 
    dlat = lat2 - lat1
    a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
 
    c = 2 * asin(sqrt(a)) 
    
    # Radius of earth in kilometers. Use 3956 for miles
    r = 6371
      
    # calculate the result
    return(c * r)


def calcDistances(cityList):
    distMatrix = np.zeros((len(cityList), len(cityList)), dtype = float)
    for row in range(0,len(cityList)):
        for col in range(0,len(cityList)):
            distMatrix[row,col] = distance(cityList[row][1], cityList[col][1], cityList[row][2], cityList[col][2])
    return distMatrix

distMat = calcDistances(cityList)
 
def selectPath(cityList, distMat, startCity):
    
    try:
        #start in the first city
        bestPath = [startingCity]
        
        #getting just the city names so we can use to index and create our path more easily
        simpleCityList = []
        for i in range(0,len(cityList)):
             simpleCityList += [cityList[i][0]]
        
        visitedInd = [simpleCityList.index(startingCity)]
    except:
        sys.exit("City not found in data, please check input and run again")
    
    
    #greedy algorithm
    allVisited = False
    totalDistance = 0
    while allVisited == False:
        if len(simpleCityList) == len(bestPath):
            break
        else:
            #find where we are
            currentInd = visitedInd[-1]
            #get the distances to other cities
            travelOptions = distMat[currentInd,:]
            #choose the next city based on min value that is not 0 (that is where we are)
            travelDistances = np.argsort(travelOptions)
            #also need to make sure we haven't visited that city yet
            #this gives us the cities we haven't visited yet
            availableCities = travelDistances[np.in1d(travelDistances[:], visitedInd, invert = True)]
            #then we go to the one with the least distance from where we are
            nextCity = availableCities[0]
            totalDistance += distMat[currentInd, nextCity]
            #add selection to bestPath
            bestPath += [simpleCityList[nextCity]]
            visitedInd += [simpleCityList.index(bestPath[-1])]
    #add the return trip back to Cleveland
    totalDistance += distMat[nextCity, visitedInd[0]]
    bestPath += [bestPath[0]]
    return bestPath, totalDistance

travelPath, totalDistance = selectPath(cityList, distMat, startingCity)

def putData(bestPath, totalDistance, filepath):
    print("now put the data back in the excel sheet")

        
        